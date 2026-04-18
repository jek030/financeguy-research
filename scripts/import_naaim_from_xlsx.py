from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_XLSX = Path(r"C:\Users\james\Downloads\USE_Data-since-Inception_2026-04-15.xlsx")
OUT_DIR = Path(r"C:\Users\james\Documents\Projects\financeguy-research")
CHUNK_SIZE = 180


def to_float(value):
    if value is None:
        return None
    return float(value)


def sql_num(value) -> str:
    if value is None:
        return "null"
    return f"{value:.2f}"


def main() -> None:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[idx["Date"]] is None:
            continue

        date_value = r[idx["Date"]]
        if isinstance(date_value, datetime):
            week_ending = date_value.date().isoformat()
        else:
            week_ending = str(date_value).split(" ")[0]

        rows.append(
            (
                week_ending,
                to_float(r[idx["NAAIM Number"]]),
                to_float(r[idx["Standard Deviation"]]),
                to_float(r[idx["Quart 1 (25% at/below)"]]),
                to_float(r[idx["Quart 2 (median)"]]),
                to_float(r[idx["Quart 3 (25% at/above)"]]),
                to_float(r[idx["Most Bullish Response"]]),
                to_float(r[idx["Most Bearish Response"]]),
            )
        )

    # Some source files contain duplicate survey dates. Keep the last seen row per
    # date so each UPSERT batch has unique conflict keys.
    deduped_by_date = {row[0]: row for row in rows}
    rows = [deduped_by_date[key] for key in sorted(deduped_by_date.keys())]

    min_date = rows[0][0] if rows else ""
    max_date = rows[-1][0] if rows else ""
    summary_text = f"count={len(rows)}\nmin_date={min_date}\nmax_date={max_date}\n"
    (OUT_DIR / ".tmp_naaim_summary.txt").write_text(summary_text, encoding="utf-8")

    for i in range(0, len(rows), CHUNK_SIZE):
        part = rows[i : i + CHUNK_SIZE]
        lines: list[str] = []

        if i == 0:
            lines.append("delete from public.market_sentiment_naaim;")

        lines.append(
            "insert into public.market_sentiment_naaim "
            "(week_ending, mean_exposure, deviation, q1, q2, q3, most_bullish, most_bearish) values"
        )

        value_lines: list[str] = []
        for t in part:
            value_lines.append(
                f"('{t[0]}', {sql_num(t[1])}, {sql_num(t[2])}, {sql_num(t[3])}, {sql_num(t[4])}, "
                f"{sql_num(t[5])}, {sql_num(t[6])}, {sql_num(t[7])})"
            )

        lines.append(",\n".join(value_lines))
        lines.append(
            "on conflict (week_ending) do update set\n"
            "  mean_exposure = excluded.mean_exposure,\n"
            "  deviation = excluded.deviation,\n"
            "  q1 = excluded.q1,\n"
            "  q2 = excluded.q2,\n"
            "  q3 = excluded.q3,\n"
            "  most_bullish = excluded.most_bullish,\n"
            "  most_bearish = excluded.most_bearish;"
        )

        part_num = (i // CHUNK_SIZE) + 1
        (OUT_DIR / f".tmp_naaim_import_{part_num:02d}.sql").write_text(
            "\n".join(lines), encoding="utf-8"
        )


if __name__ == "__main__":
    main()
