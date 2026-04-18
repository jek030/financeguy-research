from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Any

import requests
from openpyxl import load_workbook

EXPECTED_COLUMNS = {
    "date": "Date",
    "mean": "NAAIM Number",
    "deviation": "Standard Deviation",
    "q1": "Quart 1 (25% at/below)",
    "q2": "Quart 2 (median)",
    "q3": "Quart 3 (25% at/above)",
    "most_bullish": "Most Bullish Response",
    "most_bearish": "Most Bearish Response",
}


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    return float(value)


def resolve_source_bytes() -> bytes:
    source_path = os.environ.get("NAAIM_SOURCE_PATH")
    if source_path:
        with open(source_path, "rb") as file_obj:
            return file_obj.read()

    source_url = os.environ.get("NAAIM_SOURCE_URL")
    if not source_url:
        raise RuntimeError("Set NAAIM_SOURCE_URL (or NAAIM_SOURCE_PATH) before running this job.")

    response = requests.get(source_url, timeout=60)
    response.raise_for_status()
    return response.content


def extract_rows(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header_row = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_map = {str(name).strip(): index for index, name in enumerate(header_row) if name is not None}

    missing = [column for column in EXPECTED_COLUMNS.values() if column not in header_map]
    if missing:
        raise RuntimeError(f"NAAIM file is missing expected columns: {missing}")

    rows_by_date: dict[str, dict[str, Any]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        raw_date = row[header_map[EXPECTED_COLUMNS["date"]]]
        if raw_date is None:
            continue

        if isinstance(raw_date, datetime):
            week_ending = raw_date.date().isoformat()
        else:
            week_ending = str(raw_date).split(" ")[0]

        rows_by_date[week_ending] = {
            "week_ending": week_ending,
            "mean_exposure": to_float(row[header_map[EXPECTED_COLUMNS["mean"]]]),
            "deviation": to_float(row[header_map[EXPECTED_COLUMNS["deviation"]]]),
            "q1": to_float(row[header_map[EXPECTED_COLUMNS["q1"]]]),
            "q2": to_float(row[header_map[EXPECTED_COLUMNS["q2"]]]),
            "q3": to_float(row[header_map[EXPECTED_COLUMNS["q3"]]]),
            "most_bullish": to_float(row[header_map[EXPECTED_COLUMNS["most_bullish"]]]),
            "most_bearish": to_float(row[header_map[EXPECTED_COLUMNS["most_bearish"]]]),
        }

    records = [rows_by_date[key] for key in sorted(rows_by_date.keys())]
    return records


def upsert_records(records: list[dict[str, Any]]) -> None:
    supabase_url = os.environ.get("SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_role_key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required.")

    endpoint = f"{supabase_url.rstrip('/')}/rest/v1/market_sentiment_naaim?on_conflict=week_ending"
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }

    chunk_size = 250
    for start in range(0, len(records), chunk_size):
        chunk = records[start : start + chunk_size]
        response = requests.post(endpoint, headers=headers, json=chunk, timeout=60)
        response.raise_for_status()


def main() -> None:
    xlsx_bytes = resolve_source_bytes()
    records = extract_rows(xlsx_bytes)
    if not records:
        raise RuntimeError("No NAAIM rows were parsed from source file.")

    upsert_records(records)

    print(f"Upserted {len(records)} NAAIM rows into market_sentiment_naaim.")
    print(f"Range: {records[0]['week_ending']} -> {records[-1]['week_ending']}")


if __name__ == "__main__":
    main()
