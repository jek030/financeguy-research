from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_XLSX = Path(r"C:\Users\james\Downloads\USE_Data-since-Inception_2026-04-15.xlsx")
OUTPUT_SQL = Path(r"C:\Users\james\Documents\Projects\financeguy-research\.tmp_naaim_recent_import.sql")
KEEP_ROWS = 130


def num(value):
    if value is None:
        return None
    return float(value)


def sql_num(value) -> str:
    if value is None:
        return "null"
    return f"{value:.2f}"


def main() -> None:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[idx["Date"]] is None:
            continue

        d = r[idx["Date"]]
        week_ending = d.date().isoformat() if isinstance(d, datetime) else str(d).split(" ")[0]
        rows.append(
            (
                week_ending,
                num(r[idx["NAAIM Number"]]),
                num(r[idx["Standard Deviation"]]),
                num(r[idx["Quart 1 (25% at/below)"]]),
                num(r[idx["Quart 2 (median)"]]),
                num(r[idx["Quart 3 (25% at/above)"]]),
                num(r[idx["Most Bullish Response"]]),
                num(r[idx["Most Bearish Response"]]),
            )
        )

    deduped = {row[0]: row for row in rows}
    recent = [deduped[k] for k in sorted(deduped.keys())][-KEEP_ROWS:]

    lines = [
        "delete from public.market_sentiment_naaim;",
        (
            "insert into public.market_sentiment_naaim "
            "(week_ending, mean_exposure, deviation, q1, q2, q3, most_bullish, most_bearish) values"
        ),
    ]

    values = []
    for t in recent:
        values.append(
            f"('{t[0]}', {sql_num(t[1])}, {sql_num(t[2])}, {sql_num(t[3])}, "
            f"{sql_num(t[4])}, {sql_num(t[5])}, {sql_num(t[6])}, {sql_num(t[7])})"
        )
    lines.append(",\n".join(values))
    lines.append(
        "on conflict (week_ending) do update set\n"
        "  mean_exposure = excluded.mean_exposure,\n"
        "  deviation = excluded.deviation,\n"
        "  q1 = excluded.q1,\n"
        "  q2 = excluded.q2,\n"
        "  q3 = excluded.q3,\n"
        "  most_bullish = excluded.most_bullish,\n"
        "  most_bearish = excluded.most_bearish;"
    )

    OUTPUT_SQL.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
